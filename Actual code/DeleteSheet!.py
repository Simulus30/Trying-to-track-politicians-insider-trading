# Load packages
from openpyxl import load_workbook
import pandas as pd
import time

# Load the workbook
file_path = 'THIS IS AGAIN TRADES.XLSX'  # Replace with the actual path of your file
wb = load_workbook(file_path)

# Delete sheet named "Sheet"
if "Sheet" in wb.sheetnames:
    # Delete the sheet
    del wb["Sheet"]
    print("Sheet 'Sheet' has been deleted.")
else:
    print("Sheet 'Sheet' does not exist.")

# Save the workbook to apply changes
wb.save(file_path)

print(" ")
print(" ")
print(" ")
print("DONE! times Three!!!)")
print("I have killed the sheet named sheet ")
print("   O_O   ")
print(" ")
print("Running ScienceTime2.py")

# Print ".  .  ." with a delay of 1.5 seconds between each "."
for _ in range(3):
    print(".", end=" ", flush=True)  # Print on the same line
    time.sleep(1)  # Wait for 1 seconds

print()  # Move to the next line after the loop

exec(open('ScienceTime2.py').read())